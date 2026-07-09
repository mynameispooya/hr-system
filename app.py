# app.py
from flask import Flask, request, jsonify, session, send_file, send_from_directory
from functools import wraps
from datetime import datetime, timedelta
import json
import os
import uuid
import hashlib
import secrets
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import threading

app = Flask(__name__, static_folder='.', static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.permanent_session_lifetime = timedelta(hours=8)

# ============================================================================
# DATA FILE PATHS
# ============================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, 'data.json')
LOGS_FILE = os.path.join(BASE_DIR, 'logs.json')

# ============================================================================
# FILE LOCK FOR CONCURRENT ACCESS
# ============================================================================
data_lock = threading.Lock()

# ============================================================================
# DATA INITIALIZATION
# ============================================================================
def get_default_data():
    return {
        "users": {
            "admin": {
                "password": hashlib.sha256("admin123".encode()).hexdigest(),
                "role": "admin",
                "sections": [],
                "created_at": datetime.now().isoformat()
            }
        },
        "sections": {},
        "transfers": [],
        "settlements": [],
        "announcements": []
    }

def load_data():
    with data_lock:
        if not os.path.exists(DATA_FILE):
            data = get_default_data()
            save_data(data)
            return data
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            data = get_default_data()
            save_data(data)
            return data

def save_data(data):
    with data_lock:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

def load_logs():
    if not os.path.exists(LOGS_FILE):
        return []
    try:
        with open(LOGS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []

def save_logs(logs):
    with open(LOGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)

def add_log(action, user, details):
    logs = load_logs()
    log_entry = {
        "id": str(uuid.uuid4()),
        "timestamp": datetime.now().isoformat(),
        "action": action,
        "user": user,
        "details": details
    }
    logs.append(log_entry)
    # Keep only last 5000 logs
    if len(logs) > 5000:
        logs = logs[-5000:]
    save_logs(logs)

def add_announcement(message, announcement_type="info"):
    data = load_data()
    announcement = {
        "id": str(uuid.uuid4()),
        "message": message,
        "type": announcement_type,
        "timestamp": datetime.now().isoformat()
    }
    data["announcements"].insert(0, announcement)
    # Keep only last 50 announcements
    data["announcements"] = data["announcements"][:50]
    save_data(data)

# ============================================================================
# AUTH DECORATORS
# ============================================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return jsonify({"error": "لطفاً وارد شوید"}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return jsonify({"error": "لطفاً وارد شوید"}), 401
        data = load_data()
        user = data["users"].get(session['username'])
        if not user or user["role"] != "admin":
            return jsonify({"error": "دسترسی غیرمجاز"}), 403
        return f(*args, **kwargs)
    return decorated

def check_section_access(section_name):
    """Check if current user has access to the given section"""
    data = load_data()
    user = data["users"].get(session['username'])
    if not user:
        return False
    if user["role"] == "admin":
        return True
    return section_name in user.get("sections", [])

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================
def get_all_personnel_codes(data):
    """Get all personnel codes across all sections"""
    codes = set()
    for section_name, section_data in data["sections"].items():
        for code in section_data.get("employees", {}):
            codes.add(code)
    # Also include codes in pending transfers
    for transfer in data.get("transfers", []):
        if transfer.get("status") == "pending":
            codes.add(transfer["personnel_code"])
    return codes

def find_employee_by_code(data, personnel_code):
    """Find which section an employee belongs to"""
    for section_name, section_data in data["sections"].items():
        if personnel_code in section_data.get("employees", {}):
            return section_name, section_data["employees"][personnel_code]
    return None, None

# ============================================================================
# STATIC FILE ROUTES
# ============================================================================
@app.route('/')
def serve_index():
    if 'username' in session:
        return send_from_directory('.', 'index.html')
    return send_from_directory('.', 'login.html')

@app.route('/login')
def serve_login():
    return send_from_directory('.', 'login.html')

@app.route('/style.css')
def serve_css():
    return send_from_directory('.', 'style.css')

@app.route('/script.js')
def serve_js():
    return send_from_directory('.', 'script.js')

# ============================================================================
# AUTH ROUTES
# ============================================================================
@app.route('/api/login', methods=['POST'])
def login():
    data_req = request.get_json()
    if not data_req:
        return jsonify({"error": "اطلاعات ورود ارسال نشده"}), 400

    username = data_req.get('username', '').strip()
    password = data_req.get('password', '').strip()

    if not username or not password:
        return jsonify({"error": "نام کاربری و رمز عبور الزامی است"}), 400

    data = load_data()
    user = data["users"].get(username)

    if not user:
        return jsonify({"error": "نام کاربری یا رمز عبور اشتباه است"}), 401

    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    if user["password"] != hashed_password:
        return jsonify({"error": "نام کاربری یا رمز عبور اشتباه است"}), 401

    # Set session
    session.permanent = True
    session['username'] = username
    session['role'] = user["role"]
    session['sections'] = user.get("sections", [])

    add_log("ورود به سیستم", username, f"کاربر {username} وارد سیستم شد")

    return jsonify({
        "message": "ورود موفق",
        "user": {
            "username": username,
            "role": user["role"],
            "sections": user.get("sections", [])
        }
    })

@app.route('/api/logout', methods=['POST'])
@login_required
def logout():
    username = session.get('username')
    add_log("خروج از سیستم", username, f"کاربر {username} از سیستم خارج شد")
    session.clear()
    return jsonify({"message": "خروج موفق"})

@app.route('/api/me', methods=['GET'])
@login_required
def get_current_user():
    data = load_data()
    user = data["users"].get(session['username'])
    if not user:
        session.clear()
        return jsonify({"error": "کاربر یافت نشد"}), 404

    return jsonify({
        "username": session['username'],
        "role": user["role"],
        "sections": user.get("sections", [])
    })

# ============================================================================
# SECTION ROUTES
# ============================================================================
@app.route('/api/sections', methods=['GET'])
@login_required
def get_sections():
    data = load_data()
    user = data["users"].get(session['username'])
    sections = []

    for section_name, section_data in data["sections"].items():
        # Check access
        if user["role"] == "admin" or section_name in user.get("sections", []):
            employee_count = len(section_data.get("employees", {}))
            sections.append({
                "name": section_name,
                "employee_count": employee_count,
                "created_at": section_data.get("created_at", "")
            })

    return jsonify({"sections": sections})

@app.route('/api/sections', methods=['POST'])
@login_required
def create_section():
    data_req = request.get_json()
    if not data_req:
        return jsonify({"error": "اطلاعات بخش ارسال نشده"}), 400

    section_name = data_req.get('name', '').strip()
    if not section_name:
        return jsonify({"error": "نام بخش الزامی است"}), 400

    data = load_data()

    if section_name in data["sections"]:
        return jsonify({"error": "بخشی با این نام قبلاً وجود دارد"}), 409

    data["sections"][section_name] = {
        "employees": {},
        "created_at": datetime.now().isoformat()
    }
    save_data(data)

    add_log("ایجاد بخش", session['username'], f"بخش '{section_name}' ایجاد شد")
    add_announcement(f"بخش جدید '{section_name}' ایجاد شد", "info")

    return jsonify({
        "message": "بخش با موفقیت ایجاد شد",
        "section": {
            "name": section_name,
            "employee_count": 0,
            "created_at": data["sections"][section_name]["created_at"]
        }
    }), 201

@app.route('/api/sections/<section_name>', methods=['DELETE'])
@login_required
def delete_section(section_name):
    if not check_section_access(section_name):
        return jsonify({"error": "دسترسی غیرمجاز به این بخش"}), 403

    data = load_data()

    if section_name not in data["sections"]:
        return jsonify({"error": "بخش مورد نظر یافت نشد"}), 404

    employee_count = len(data["sections"][section_name].get("employees", {}))
    del data["sections"][section_name]

    # Remove related transfers
    data["transfers"] = [t for t in data["transfers"]
                         if t.get("from_section") != section_name and t.get("to_section") != section_name]

    save_data(data)

    add_log("حذف بخش", session['username'],
            f"بخش '{section_name}' با {employee_count} کارمند حذف شد")

    return jsonify({"message": "بخش با موفقیت حذف شد"})

# ============================================================================
# EMPLOYEE ROUTES
# ============================================================================
@app.route('/api/sections/<section_name>/employees', methods=['GET'])
@login_required
def get_employees(section_name):
    if not check_section_access(section_name):
        return jsonify({"error": "دسترسی غیرمجاز به این بخش"}), 403

    data = load_data()

    if section_name not in data["sections"]:
        return jsonify({"error": "بخش مورد نظر یافت نشد"}), 404

    employees = data["sections"][section_name].get("employees", {})
    employee_list = list(employees.values())

    return jsonify({"employees": employee_list, "section": section_name})

@app.route('/api/sections/<section_name>/employees', methods=['POST'])
@login_required
def add_employee(section_name):
    if not check_section_access(section_name):
        return jsonify({"error": "دسترسی غیرمجاز به این بخش"}), 403

    data_req = request.get_json()
    if not data_req:
        return jsonify({"error": "اطلاعات کارمند ارسال نشده"}), 400

    first_name = data_req.get('first_name', '').strip()
    last_name = data_req.get('last_name', '').strip()
    personnel_code = str(data_req.get('personnel_code', '')).strip()

    if not first_name or not last_name or not personnel_code:
        return jsonify({"error": "نام، نام خانوادگی و کد پرسنلی الزامی است"}), 400

    data = load_data()

    if section_name not in data["sections"]:
        return jsonify({"error": "بخش مورد نظر یافت نشد"}), 404

    # Check uniqueness across ALL sections
    all_codes = get_all_personnel_codes(data)
    if personnel_code in all_codes:
        return jsonify({"error": f"کد پرسنلی '{personnel_code}' قبلاً در سیستم ثبت شده است"}), 409

    employee = {
        "personnel_code": personnel_code,
        "first_name": first_name,
        "last_name": last_name,
        "added_at": datetime.now().isoformat(),
        "added_by": session['username']
    }

    data["sections"][section_name]["employees"][personnel_code] = employee
    save_data(data)

    full_name = f"{first_name} {last_name}"
    add_log("افزودن کارمند", session['username'],
            f"{full_name} (کد: {personnel_code}) به بخش '{section_name}' اضافه شد")

    return jsonify({
        "message": "کارمند با موفقیت اضافه شد",
        "employee": employee
    }), 201

@app.route('/api/sections/<section_name>/employees/<personnel_code>', methods=['DELETE'])
@login_required
def delete_employee(section_name, personnel_code):
    if not check_section_access(section_name):
        return jsonify({"error": "دسترسی غیرمجاز به این بخش"}), 403

    data = load_data()

    if section_name not in data["sections"]:
        return jsonify({"error": "بخش مورد نظر یافت نشد"}), 404

    employees = data["sections"][section_name].get("employees", {})
    if personnel_code not in employees:
        return jsonify({"error": "کارمند مورد نظر یافت نشد"}), 404

    employee = employees[personnel_code]
    del data["sections"][section_name]["employees"][personnel_code]

    # Remove related transfers
    data["transfers"] = [t for t in data["transfers"]
                         if t.get("personnel_code") != personnel_code]

    save_data(data)

    full_name = f"{employee['first_name']} {employee['last_name']}"
    add_log("حذف کارمند", session['username'],
            f"{full_name} (کد: {personnel_code}) از بخش '{section_name}' حذف شد")

    return jsonify({"message": "کارمند با موفقیت حذف شد"})

# ============================================================================
# EXCEL UPLOAD / DOWNLOAD
# ============================================================================
@app.route('/api/sections/<section_name>/upload', methods=['POST'])
@login_required
def upload_excel(section_name):
    if not check_section_access(section_name):
        return jsonify({"error": "دسترسی غیرمجاز به این بخش"}), 403

    if 'file' not in request.files:
        return jsonify({"error": "فایلی ارسال نشده است"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "فایلی انتخاب نشده است"}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "فقط فایل‌های اکسل (.xlsx, .xls) مجاز هستند"}), 400

    data = load_data()
    if section_name not in data["sections"]:
        return jsonify({"error": "بخش مورد نظر یافت نشد"}), 404

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active

        added = []
        skipped = []
        errors = []
        all_codes = get_all_personnel_codes(data)

        # Find header row
        header_row = None
        first_name_col = None
        last_name_col = None
        code_col = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_values = [str(cell).strip().lower() if cell else '' for cell in row]
            # Look for Persian or English headers
            for col_idx, val in enumerate(row_values):
                if any(keyword in val for keyword in ['نام', 'first', 'name']):
                    if first_name_col is None:
                        first_name_col = col_idx
                if any(keyword in val for keyword in ['خانوادگی', 'last', 'family']):
                    last_name_col = col_idx
                if any(keyword in val for keyword in ['کد پرسنلی', 'کدپرسنلی', 'personnel', 'code', 'کد']):
                    code_col = col_idx

            if first_name_col is not None and last_name_col is not None and code_col is not None:
                header_row = row_idx
                break

        if header_row is None:
            # Try to guess columns: assume first 3 columns are first_name, last_name, code
            header_row = 1
            first_name_col = 0
            last_name_col = 1
            code_col = 2

        # Read data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            try:
                first_name = str(row[first_name_col] or '').strip()
                last_name = str(row[last_name_col] or '').strip()
                personnel_code = str(row[code_col] or '').strip()

                if not first_name or not last_name or not personnel_code:
                    errors.append(f"ردیف ناقص: {row}")
                    continue

                # Remove .0 from float codes
                if personnel_code.endswith('.0'):
                    personnel_code = personnel_code[:-2]

                if personnel_code in all_codes:
                    skipped.append(f"{first_name} {last_name} (کد: {personnel_code}) - تکراری")
                    continue

                employee = {
                    "personnel_code": personnel_code,
                    "first_name": first_name,
                    "last_name": last_name,
                    "added_at": datetime.now().isoformat(),
                    "added_by": session['username']
                }

                data["sections"][section_name]["employees"][personnel_code] = employee
                all_codes.add(personnel_code)
                added.append(f"{first_name} {last_name} (کد: {personnel_code})")

            except (IndexError, ValueError) as e:
                errors.append(f"خطا در خواندن ردیف: {str(e)}")
                continue

        save_data(data)
        wb.close()

        if added:
            add_log("آپلود اکسل", session['username'],
                    f"{len(added)} کارمند از فایل اکسل به بخش '{section_name}' اضافه شد")

        return jsonify({
            "message": f"{len(added)} کارمند اضافه شد",
            "added_count": len(added),
            "skipped_count": len(skipped),
            "error_count": len(errors),
            "added": added[:20],  # Limit response size
            "skipped": skipped[:20],
            "errors": errors[:10]
        })

    except Exception as e:
        return jsonify({"error": f"خطا در خواندن فایل اکسل: {str(e)}"}), 400

@app.route('/api/sections/<section_name>/export', methods=['GET'])
@login_required
def export_excel(section_name):
    if not check_section_access(section_name):
        return jsonify({"error": "دسترسی غیرمجاز به این بخش"}), 403

    data = load_data()
    if section_name not in data["sections"]:
        return jsonify({"error": "بخش مورد نظر یافت نشد"}), 404

    employees = data["sections"][section_name].get("employees", {})

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = section_name[:31]  # Excel sheet name max 31 chars

    # RTL support
    ws.sheet_view.rightToLeft = True

    # Header style
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ["ردیف", "کد پرسنلی", "نام", "نام خانوادگی"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data
    for row_idx, (code, emp) in enumerate(employees.items(), 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=emp["personnel_code"])
        ws.cell(row=row_idx, column=3, value=emp["first_name"])
        ws.cell(row=row_idx, column=4, value=emp["last_name"])

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    add_log("خروجی اکسل", session['username'],
            f"خروجی اکسل از بخش '{section_name}' با {len(employees)} کارمند گرفته شد")

    filename = f"{section_name}_employees.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ============================================================================
# TRANSFER ROUTES
# ============================================================================
@app.route('/api/transfers', methods=['POST'])
@login_required
def request_transfer():
    data_req = request.get_json()
    if not data_req:
        return jsonify({"error": "اطلاعات انتقال ارسال نشده"}), 400

    personnel_code = str(data_req.get('personnel_code', '')).strip()
    to_section = data_req.get('to_section', '').strip()

    if not personnel_code or not to_section:
        return jsonify({"error": "کد پرسنلی و بخش مقصد الزامی است"}), 400

    data = load_data()

    # Check if to_section exists
    if to_section not in data["sections"]:
        return jsonify({"error": "بخش مقصد یافت نشد"}), 404

    # Check access to from_section
    from_section, employee = find_employee_by_code(data, personnel_code)
    if not from_section:
        return jsonify({"error": "کارمند با این کد پرسنلی یافت نشد"}), 404

    if not check_section_access(from_section):
        return jsonify({"error": "دسترسی غیرمجاز به بخش مبدأ"}), 403

    if from_section == to_section:
        return jsonify({"error": "بخش مبدأ و مقصد یکسان هستند"}), 400

    # Check if there's already a pending transfer for this employee
    for transfer in data["transfers"]:
        if transfer["personnel_code"] == personnel_code and transfer["status"] == "pending":
            return jsonify({"error": "یک درخواست انتقال برای این کارمند قبلاً ثبت شده و در انتظار تأیید است"}), 409

    # Create transfer request
    transfer = {
        "id": str(uuid.uuid4()),
        "personnel_code": personnel_code,
        "first_name": employee["first_name"],
        "last_name": employee["last_name"],
        "from_section": from_section,
        "to_section": to_section,
        "status": "pending",
        "requested_by": session['username'],
        "requested_at": datetime.now().isoformat(),
        "approved_by": None,
        "approved_at": None,
        "rejected_by": None,
        "rejected_at": None,
        "rejection_reason": ""
    }

    data["transfers"].append(transfer)
    save_data(data)

    full_name = f"{employee['first_name']} {employee['last_name']}"
    add_log("درخواست انتقال", session['username'],
            f"درخواست انتقال {full_name} (کد: {personnel_code}) از '{from_section}' به '{to_section}'")

    return jsonify({
        "message": "درخواست انتقال با موفقیت ثبت شد",
        "transfer": transfer
    }), 201

@app.route('/api/transfers/pending', methods=['GET'])
@login_required
def get_pending_transfers():
    data = load_data()
    user = data["users"].get(session['username'])

    pending = []
    for transfer in data["transfers"]:
        if transfer["status"] != "pending":
            continue

        # User can see if they have access to from_section or to_section
        can_see = False
        if user["role"] == "admin":
            can_see = True
        elif transfer["from_section"] in user.get("sections", []):
            can_see = True
        elif transfer["to_section"] in user.get("sections", []):
            can_see = True

        if can_see:
            pending.append(transfer)

    return jsonify({"transfers": pending})

@app.route('/api/transfers/my-pending', methods=['GET'])
@login_required
def get_my_pending_transfers():
    """Get pending transfers where current user's section is the destination"""
    data = load_data()
    user = data["users"].get(session['username'])

    my_pending = []
    for transfer in data["transfers"]:
        if transfer["status"] != "pending":
            continue

        # User can approve if they have access to the destination section
        can_approve = False
        if user["role"] == "admin":
            can_approve = True
        elif transfer["to_section"] in user.get("sections", []):
            can_approve = True

        if can_approve:
            my_pending.append(transfer)

    return jsonify({"transfers": my_pending})

@app.route('/api/transfers/<transfer_id>/approve', methods=['POST'])
@login_required
def approve_transfer(transfer_id):
    data = load_data()
    user = data["users"].get(session['username'])

    # Find transfer
    transfer = None
    for t in data["transfers"]:
        if t["id"] == transfer_id:
            transfer = t
            break

    if not transfer:
        return jsonify({"error": "درخواست انتقال یافت نشد"}), 404

    if transfer["status"] != "pending":
        return jsonify({"error": "این درخواست قبلاً پردازش شده"}), 400

    # Check if user can approve (has access to destination section)
    if user["role"] != "admin" and transfer["to_section"] not in user.get("sections", []):
        return jsonify({"error": "دسترسی غیرمجاز برای تأیید این انتقال"}), 403

    # Check if employee still exists in source section
    from_section = transfer["from_section"]
    personnel_code = transfer["personnel_code"]

    if from_section not in data["sections"]:
        return jsonify({"error": "بخش مبدأ دیگر وجود ندارد"}), 404

    if personnel_code not in data["sections"][from_section].get("employees", {}):
        return jsonify({"error": "کارمند دیگر در بخش مبدأ وجود ندارد"}), 404

    # Perform transfer
    employee = data["sections"][from_section]["employees"].pop(personnel_code)

    # Add to destination section
    if transfer["to_section"] not in data["sections"]:
        data["sections"][transfer["to_section"]] = {
            "employees": {},
            "created_at": datetime.now().isoformat()
        }

    employee["transferred_at"] = datetime.now().isoformat()
    employee["transferred_by"] = session['username']
    data["sections"][transfer["to_section"]]["employees"][personnel_code] = employee

    # Update transfer status
    transfer["status"] = "approved"
    transfer["approved_by"] = session['username']
    transfer["approved_at"] = datetime.now().isoformat()

    save_data(data)

    full_name = f"{transfer['first_name']} {transfer['last_name']}"
    add_log("تأیید انتقال", session['username'],
            f"انتقال {full_name} (کد: {personnel_code}) از '{from_section}' به '{transfer['to_section']}' تأیید شد")

    add_announcement(
        f"✅ {full_name} (کد پرسنلی: {personnel_code}) از بخش '{from_section}' به بخش '{transfer['to_section']}' منتقل شد",
        "transfer"
    )

    return jsonify({
        "message": "انتقال با موفقیت تأیید و انجام شد",
        "transfer": transfer
    })

@app.route('/api/transfers/<transfer_id>/reject', methods=['POST'])
@login_required
def reject_transfer(transfer_id):
    data_req = request.get_json() or {}
    reason = data_req.get('reason', '').strip()

    data = load_data()
    user = data["users"].get(session['username'])

    # Find transfer
    transfer = None
    for t in data["transfers"]:
        if t["id"] == transfer_id:
            transfer = t
            break

    if not transfer:
        return jsonify({"error": "درخواست انتقال یافت نشد"}), 404

    if transfer["status"] != "pending":
        return jsonify({"error": "این درخواست قبلاً پردازش شده"}), 400

    # Check if user can reject
    if user["role"] != "admin" and transfer["to_section"] not in user.get("sections", []):
        return jsonify({"error": "دسترسی غیرمجاز برای رد این انتقال"}), 403

    transfer["status"] = "rejected"
    transfer["rejected_by"] = session['username']
    transfer["rejected_at"] = datetime.now().isoformat()
    transfer["rejection_reason"] = reason

    save_data(data)

    full_name = f"{transfer['first_name']} {transfer['last_name']}"
    add_log("رد انتقال", session['username'],
            f"انتقال {full_name} (کد: {personnel_code}) از '{transfer['from_section']}' به '{transfer['to_section']}' رد شد. دلیل: {reason or 'نامشخص'}")

    return jsonify({
        "message": "درخواست انتقال رد شد",
        "transfer": transfer
    })

@app.route('/api/transfers/history', methods=['GET'])
@login_required
def get_transfer_history():
    data = load_data()
    user = data["users"].get(session['username'])

    history = []
    for transfer in data["transfers"]:
        if transfer["status"] == "pending":
            continue

        can_see = False
        if user["role"] == "admin":
            can_see = True
        elif transfer["from_section"] in user.get("sections", []):
            can_see = True
        elif transfer["to_section"] in user.get("sections", []):
            can_see = True

        if can_see:
            history.append(transfer)

    # Sort by most recent
    history.sort(key=lambda x: x.get("approved_at") or x.get("rejected_at") or "", reverse=True)

    return jsonify({"transfers": history})

# ============================================================================
# SETTLEMENT ROUTES
# ============================================================================
@app.route('/api/settlements', methods=['POST'])
@login_required
def request_settlement():
    data_req = request.get_json()
    if not data_req:
        return jsonify({"error": "اطلاعات تسویه ارسال نشده"}), 400

    personnel_code = str(data_req.get('personnel_code', '')).strip()

    if not personnel_code:
        return jsonify({"error": "کد پرسنلی الزامی است"}), 400

    data = load_data()

    # Find employee
    from_section, employee = find_employee_by_code(data, personnel_code)
    if not from_section:
        return jsonify({"error": "کارمند با این کد پرسنلی یافت نشد"}), 404

    if not check_section_access(from_section):
        return jsonify({"error": "دسترسی غیرمجاز به بخش این کارمند"}), 403

    # Check if already in settlement
    for s in data["settlements"]:
        if s["personnel_code"] == personnel_code:
            return jsonify({"error": "این کارمند قبلاً تسویه حساب شده است"}), 409

    # Remove from section
    del data["sections"][from_section]["employees"][personnel_code]

    # Add to settlements
    settlement = {
        "id": str(uuid.uuid4()),
        "personnel_code": personnel_code,
        "first_name": employee["first_name"],
        "last_name": employee["last_name"],
        "from_section": from_section,
        "settled_by": session['username'],
        "settled_at": datetime.now().isoformat(),
        "notes": data_req.get('notes', '')
    }

    data["settlements"].append(settlement)

    # Remove related transfers
    data["transfers"] = [t for t in data["transfers"]
                         if t.get("personnel_code") != personnel_code]

    save_data(data)

    full_name = f"{employee['first_name']} {employee['last_name']}"
    add_log("تسویه حساب", session['username'],
            f"{full_name} (کد: {personnel_code}) از بخش '{from_section}' تسویه حساب شد")

    add_announcement(
        f"🔴 {full_name} (کد پرسنلی: {personnel_code}) از بخش '{from_section}' تسویه حساب شد",
        "settlement"
    )

    return jsonify({
        "message": "تسویه حساب با موفقیت انجام شد",
        "settlement": settlement
    })

@app.route('/api/settlements', methods=['GET'])
@login_required
def get_settlements():
    data = load_data()
    user = data["users"].get(session['username'])

    settlements = []
    for s in data["settlements"]:
        can_see = False
        if user["role"] == "admin":
            can_see = True
        elif s["from_section"] in user.get("sections", []):
            can_see = True

        if can_see:
            settlements.append(s)

    settlements.sort(key=lambda x: x.get("settled_at", ""), reverse=True)

    return jsonify({"settlements": settlements})

# ============================================================================
# USER MANAGEMENT ROUTES (Admin Only)
# ============================================================================
@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    data = load_data()
    users = []
    for username, user_data in data["users"].items():
        users.append({
            "username": username,
            "role": user_data["role"],
            "sections": user_data.get("sections", []),
            "created_at": user_data.get("created_at", "")
        })
    return jsonify({"users": users})

@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    data_req = request.get_json()
    if not data_req:
        return jsonify({"error": "اطلاعات کاربر ارسال نشده"}), 400

    username = data_req.get('username', '').strip()
    password = data_req.get('password', '').strip()
    role = data_req.get('role', 'user').strip()
    sections = data_req.get('sections', [])

    if not username or not password:
        return jsonify({"error": "نام کاربری و رمز عبور الزامی است"}), 400

    if len(password) < 4:
        return jsonify({"error": "رمز عبور باید حداقل ۴ کاراکتر باشد"}), 400

    if role not in ['admin', 'user']:
        return jsonify({"error": "نقش باید 'admin' یا 'user' باشد"}), 400

    data = load_data()

    if username in data["users"]:
        return jsonify({"error": "این نام کاربری قبلاً ثبت شده است"}), 409

    # Validate sections exist
    for section_name in sections:
        if section_name not in data["sections"]:
            return jsonify({"error": f"بخش '{section_name}' وجود ندارد"}), 400

    user = {
        "password": hashlib.sha256(password.encode()).hexdigest(),
        "role": role,
        "sections": sections,
        "created_at": datetime.now().isoformat()
    }

    data["users"][username] = user
    save_data(data)

    add_log("ایجاد کاربر", session['username'],
            f"کاربر '{username}' با نقش '{role}' ایجاد شد")

    return jsonify({
        "message": "کاربر با موفقیت ایجاد شد",
        "user": {
            "username": username,
            "role": role,
            "sections": sections,
            "created_at": user["created_at"]
        }
    }), 201

@app.route('/api/users/<username>', methods=['PUT'])
@admin_required
def update_user(username):
    data_req = request.get_json()
    if not data_req:
        return jsonify({"error": "اطلاعات ارسال نشده"}), 400

    data = load_data()

    if username not in data["users"]:
        return jsonify({"error": "کاربر یافت نشد"}), 404

    user = data["users"][username]

    if 'password' in data_req and data_req['password']:
        if len(data_req['password']) < 4:
            return jsonify({"error": "رمز عبور باید حداقل ۴ کاراکتر باشد"}), 400
        user["password"] = hashlib.sha256(data_req['password'].encode()).hexdigest()

    if 'role' in data_req:
        if data_req['role'] not in ['admin', 'user']:
            return jsonify({"error": "نقش نامعتبر"}), 400
        user["role"] = data_req['role']

    if 'sections' in data_req:
        for section_name in data_req['sections']:
            if section_name not in data["sections"]:
                return jsonify({"error": f"بخش '{section_name}' وجود ندارد"}), 400
        user["sections"] = data_req['sections']

    save_data(data)

    add_log("ویرایش کاربر", session['username'],
            f"اطلاعات کاربر '{username}' ویرایش شد")

    return jsonify({
        "message": "کاربر با موفقیت ویرایش شد",
        "user": {
            "username": username,
            "role": user["role"],
            "sections": user.get("sections", [])
        }
    })

@app.route('/api/users/<username>', methods=['DELETE'])
@admin_required
def delete_user(username):
    if username == session['username']:
        return jsonify({"error": "نمی‌توانید کاربر فعلی را حذف کنید"}), 400

    data = load_data()

    if username not in data["users"]:
        return jsonify({"error": "کاربر یافت نشد"}), 404

    if username == "admin":
        return jsonify({"error": "نمی‌توان کاربر admin را حذف کرد"}), 400

    del data["users"][username]
    save_data(data)

    add_log("حذف کاربر", session['username'],
            f"کاربر '{username}' حذف شد")

    return jsonify({"message": "کاربر با موفقیت حذف شد"})

# ============================================================================
# LOGS ROUTE (Admin Only)
# ============================================================================
@app.route('/api/logs', methods=['GET'])
@admin_required
def get_logs():
    logs = load_logs()

    # Optional filtering
    action_filter = request.args.get('action', '').strip()
    user_filter = request.args.get('user', '').strip()

    filtered = logs
    if action_filter:
        filtered = [l for l in filtered if action_filter in l.get("action", "")]
    if user_filter:
        filtered = [l for l in filtered if user_filter in l.get("user", "")]

    # Sort by most recent first
    filtered.sort(key=lambda x: x.get("timestamp", ""), reverse=True)

    # Pagination
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 100))
    start = (page - 1) * per_page
    end = start + per_page

    return jsonify({
        "logs": filtered[start:end],
        "total": len(filtered),
        "page": page,
        "per_page": per_page
    })

@app.route('/api/logs/export', methods=['GET'])
@admin_required
def export_logs():
    logs = load_logs()
    logs.sort(key=lambda x: x.get("timestamp", ""), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Logs"
    ws.sheet_view.rightToLeft = True

    # Headers
    headers = ["ردیف", "تاریخ و زمان", "کاربر", "عملیات", "جزئیات"]
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=log.get("timestamp", ""))
        ws.cell(row=row_idx, column=3, value=log.get("user", ""))
        ws.cell(row=row_idx, column=4, value=log.get("action", ""))
        ws.cell(row=row_idx, column=5, value=log.get("details", ""))

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 60

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name="activity_logs.xlsx"
    )

# ============================================================================
# ANNOUNCEMENTS ROUTE
# ============================================================================
@app.route('/api/announcements', methods=['GET'])
@login_required
def get_announcements():
    data = load_data()
    announcements = data.get("announcements", [])
    return jsonify({"announcements": announcements})

# ============================================================================
# CHANGE PASSWORD
# ============================================================================
@app.route('/api/change-password', methods=['POST'])
@login_required
def change_password():
    data_req = request.get_json()
    if not data_req:
        return jsonify({"error": "اطلاعات ارسال نشده"}), 400

    current_password = data_req.get('current_password', '').strip()
    new_password = data_req.get('new_password', '').strip()

    if not current_password or not new_password:
        return jsonify({"error": "رمز فعلی و رمز جدید الزامی است"}), 400

    if len(new_password) < 4:
        return jsonify({"error": "رمز جدید باید حداقل ۴ کاراکتر باشد"}), 400

    data = load_data()
    user = data["users"].get(session['username'])

    if not user:
        return jsonify({"error": "کاربر یافت نشد"}), 404

    current_hash = hashlib.sha256(current_password.encode()).hexdigest()
    if user["password"] != current_hash:
        return jsonify({"error": "رمز فعلی اشتباه است"}), 401

    user["password"] = hashlib.sha256(new_password.encode()).hexdigest()
    save_data(data)

    add_log("تغییر رمز عبور", session['username'],
            f"رمز عبور کاربر {session['username']} تغییر کرد")

    return jsonify({"message": "رمز عبور با موفقیت تغییر کرد"})

# ============================================================================
# ERROR HANDLERS
# ============================================================================
@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "صفحه مورد نظر یافت نشد"}), 404

@app.errorhandler(500)
def internal_error(e):
    return jsonify({"error": "خطای داخلی سرور"}), 500

# ============================================================================
# MAIN ENTRY POINT
# ============================================================================
if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
